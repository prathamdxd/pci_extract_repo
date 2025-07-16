from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from openpyxl import load_workbook
from typing import List, Dict
from pydantic import BaseModel
import io

app = FastAPI()

# Configure CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class RoadData(BaseModel):
    road_name: str
    pcivalue_2019: int
    pcivalue_2021: int

@app.post("/upload_excel/")
async def upload_excel(file: UploadFile = File(...)):
    try:
        # Read file content
        contents = await file.read()
        
        # Load workbook
        wb = load_workbook(filename=io.BytesIO(contents))
        sheet = wb.active
        
        # Find column indices
        headers = [cell.value for cell in sheet[1]]
        col_map = {}
        
        required_columns = {
            "road_name": str,
            "pcivalue_2019": int,
            "pcivalue_2021": int
        }
        
        # Case-insensitive header matching
        for col in required_columns:
            found = False
            for idx, header in enumerate(headers):
                if str(header).lower().strip() == col.lower():
                    col_map[col] = idx
                    found = True
                    break
            if not found:
                raise HTTPException(
                    status_code=400,
                    detail=f"Missing required column: {col}"
                )
        
        # Process rows
        results = []
        validation_errors = []
        
        for row in sheet.iter_rows(min_row=2):  # Skip header
            try:
                data = {
                    "road_name": str(row[col_map["road_name"]].value).strip(),
                    "pcivalue_2019": int(row[col_map["pcivalue_2019"]].value),
                    "pcivalue_2021": int(row[col_map["pcivalue_2021"]].value)
                }
                
                # Validate PCI values
                if not (1 <= data["pcivalue_2019"] <= 5):
                    validation_errors.append(
                        f"Row {row[0].row}: pcivalue_2019 must be 1-5, got {data['pcivalue_2019']}"
                    )
                if not (1 <= data["pcivalue_2021"] <= 5):
                    validation_errors.append(
                        f"Row {row[0].row}: pcivalue_2021 must be 1-5, got {data['pcivalue_2021']}"
                    )
                if not data["road_name"]:
                    validation_errors.append(
                        f"Row {row[0].row}: road_name cannot be empty"
                    )
                
                results.append(data)
            except (ValueError, TypeError):
                validation_errors.append(
                    f"Row {row[0].row}: Invalid data types"
                )
            except Exception as e:
                validation_errors.append(
                    f"Row {row[0].row}: Error processing row - {str(e)}"
                )
        
        if validation_errors:
            raise HTTPException(
                status_code=400,
                detail="Validation errors:\n" + "\n".join(validation_errors)
            )
        
        return results
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error processing file: {str(e)}"
        )

@app.get("/health")
async def health_check():
    return {"status": "ok"}
